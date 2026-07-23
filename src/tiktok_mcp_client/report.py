from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any


DATE_PRESETS: dict[str, tuple[str, str] | None] = {
    "today": ("${today}", "${today}"),
    "yesterday": ("${yesterday}", "${yesterday}"),
    "last_7_days": ("${week_ago}", "${today}"),
    "last_14_days": ("${days_ago_14}", "${today}"),
    "last_30_days": ("${days_ago_30}", "${today}"),
    "this_month": ("${month_start}", "${today}"),
    "lifetime": None,
}


def resolve_date_preset(
    preset: str,
    start_date: str | None = None,
    end_date: str | None = None,
) -> tuple[str | None, str | None, bool]:
    """Return (start_date, end_date, query_lifetime)."""
    key = preset.strip().lower().replace("-", "_").replace(" ", "_")
    if key in {"day", "single_day", "one_day", "date"}:
        day = start_date or end_date
        if not day:
            raise ValueError("单日查询需要提供日期，例如 --date 2026-07-22")
        return day, day, False
    if key in {"custom", "range"}:
        if start_date and not end_date:
            # 只给开始日时按单日处理
            return start_date, start_date, False
        if not start_date or not end_date:
            raise ValueError("自定义时间范围需要同时提供 start_date 和 end_date")
        return start_date, end_date, False
    if key in {"lifetime", "all", "full", "query_lifetime"}:
        return None, None, True
    if key not in DATE_PRESETS:
        raise ValueError(
            f"未知时间预设: {preset}。"
            f"可选: {', '.join(DATE_PRESETS)} / day / custom / lifetime"
        )
    pair = DATE_PRESETS[key]
    assert pair is not None
    return pair[0], pair[1], False


def to_number(value: Any) -> float:
    if value is None or value == "" or value == "-":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").replace("%", "").strip()
    try:
        return float(text)
    except ValueError:
        return 0.0


def get_metric(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in row:
            return row[key]
        nested = row.get("metrics")
        if isinstance(nested, dict) and key in nested:
            return nested[key]
        info = row.get("info")
        if isinstance(info, dict) and key in info:
            return info[key]
        dotted = f"metrics.{key}"
        if dotted in row:
            return row[dotted]
        dotted_info = f"info.{key}"
        if dotted_info in row:
            return row[dotted_info]
    return None


def build_summary(rows: list[dict[str, Any]]) -> dict[str, Any]:
    totals = {
        "row_count": len(rows),
        "spend": 0.0,
        "cash_spend": 0.0,
        "voucher_spend": 0.0,
        "impressions": 0.0,
        "clicks": 0.0,
        "engagements": 0.0,
        "conversion": 0.0,
        "native_growth_ad_revenue_value_d0": 0.0,
        "purchase_value": 0.0,
        "rows_with_spend": 0,
        "rows_with_conversion": 0,
    }
    by_advertiser: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "advertiser_id": "",
            "advertiser_name": "",
            "row_count": 0,
            "spend": 0.0,
            "cash_spend": 0.0,
            "voucher_spend": 0.0,
            "impressions": 0.0,
            "clicks": 0.0,
            "engagements": 0.0,
            "conversion": 0.0,
            "native_growth_ad_revenue_value_d0": 0.0,
            "purchase_value": 0.0,
        }
    )

    for row in rows:
        spend = to_number(get_metric(row, "spend"))
        cash_spend = to_number(get_metric(row, "cash_spend"))
        voucher_spend = to_number(get_metric(row, "voucher_spend"))
        impressions = to_number(get_metric(row, "impressions"))
        engagements = to_number(get_metric(row, "engagements", "clicks"))
        clicks = to_number(get_metric(row, "clicks", "engagements"))
        conversion = to_number(get_metric(row, "conversion"))
        native_d0 = to_number(
            get_metric(row, "native_growth_ad_revenue_value_d0")
        )
        purchase = to_number(
            get_metric(
                row,
                "native_growth_ad_revenue_value_d0",
                "total_purchase_value",
                "value_per_complete_payment",
            )
        )
        totals["spend"] += spend
        totals["cash_spend"] += cash_spend
        totals["voucher_spend"] += voucher_spend
        totals["impressions"] += impressions
        totals["clicks"] += clicks
        totals["engagements"] += engagements
        totals["conversion"] += conversion
        totals["native_growth_ad_revenue_value_d0"] += native_d0
        totals["purchase_value"] += purchase
        if spend > 0:
            totals["rows_with_spend"] += 1
        if conversion > 0:
            totals["rows_with_conversion"] += 1

        advertiser_id = str(row.get("advertiser_id") or "")
        advertiser_name = str(row.get("advertiser_name") or "")
        bucket = by_advertiser[advertiser_id or advertiser_name or "unknown"]
        bucket["advertiser_id"] = advertiser_id
        bucket["advertiser_name"] = advertiser_name
        bucket["row_count"] += 1
        bucket["spend"] += spend
        bucket["cash_spend"] += cash_spend
        bucket["voucher_spend"] += voucher_spend
        bucket["impressions"] += impressions
        bucket["clicks"] += clicks
        bucket["engagements"] += engagements
        bucket["conversion"] += conversion
        bucket["native_growth_ad_revenue_value_d0"] += native_d0
        bucket["purchase_value"] += purchase

    click_base = totals["engagements"] or totals["clicks"]
    totals["ctr"] = (
        round(click_base / totals["impressions"] * 100, 4)
        if totals["impressions"]
        else 0.0
    )
    totals["cpc"] = (
        round(totals["spend"] / click_base, 4) if click_base else 0.0
    )
    totals["cpm"] = (
        round(totals["spend"] / totals["impressions"] * 1000, 4)
        if totals["impressions"]
        else 0.0
    )
    totals["cpa"] = (
        round(totals["spend"] / totals["conversion"], 4)
        if totals["conversion"]
        else None
    )
    totals["roas"] = (
        round(totals["purchase_value"] / totals["spend"], 4)
        if totals["spend"] and totals["purchase_value"]
        else 0.0
    )
    totals["native_growth_ad_revenue_roas_d0"] = (
        round(
            totals["native_growth_ad_revenue_value_d0"] / totals["spend"], 4
        )
        if totals["spend"] and totals["native_growth_ad_revenue_value_d0"]
        else 0.0
    )
    for key in (
        "spend",
        "cash_spend",
        "voucher_spend",
        "purchase_value",
        "native_growth_ad_revenue_value_d0",
    ):
        totals[key] = round(totals[key], 2)

    advertiser_rows = []
    for item in by_advertiser.values():
        item["spend"] = round(item["spend"], 2)
        item["cash_spend"] = round(item["cash_spend"], 2)
        item["voucher_spend"] = round(item["voucher_spend"], 2)
        item["purchase_value"] = round(item["purchase_value"], 2)
        item["native_growth_ad_revenue_value_d0"] = round(
            item["native_growth_ad_revenue_value_d0"], 2
        )
        click_base = item["engagements"] or item["clicks"]
        item["ctr"] = (
            round(click_base / item["impressions"] * 100, 4)
            if item["impressions"]
            else 0.0
        )
        item["cpc"] = (
            round(item["spend"] / click_base, 4) if click_base else 0.0
        )
        item["cpa"] = (
            round(item["spend"] / item["conversion"], 4)
            if item["conversion"]
            else None
        )
        item["native_growth_ad_revenue_roas_d0"] = (
            round(
                item["native_growth_ad_revenue_value_d0"] / item["spend"], 4
            )
            if item["spend"] and item["native_growth_ad_revenue_value_d0"]
            else 0.0
        )
        advertiser_rows.append(item)
    advertiser_rows.sort(key=lambda item: item["spend"], reverse=True)

    return {
        "totals": totals,
        "by_advertiser": advertiser_rows,
    }


def filter_rows_with_activity(
    rows: list[dict[str, Any]],
    *,
    only_spend: bool = False,
) -> list[dict[str, Any]]:
    filtered: list[dict[str, Any]] = []
    for row in rows:
        spend = to_number(get_metric(row, "spend"))
        impressions = to_number(get_metric(row, "impressions"))
        clicks = to_number(get_metric(row, "clicks", "engagements"))
        conversion = to_number(get_metric(row, "conversion"))
        if only_spend:
            if spend > 0:
                filtered.append(row)
        elif spend > 0 or impressions > 0 or clicks > 0 or conversion > 0:
            filtered.append(row)
    return filtered


def write_excel(
    path: Path,
    *,
    detail_rows: list[dict[str, Any]],
    summary: dict[str, Any],
    account_summaries: list[dict[str, Any]],
    errors: list[dict[str, Any]],
    meta: dict[str, Any],
    flatten_row,
) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as error:
        raise RuntimeError(
            "缺少 openpyxl，请先执行: pip install openpyxl"
        ) from error

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()

    overview = workbook.active
    overview.title = "总览"
    overview.append(["字段", "值"])
    for key, value in {
        **meta,
        **{f"totals.{k}": v for k, v in summary.get("totals", {}).items()},
    }.items():
        overview.append([key, value if value is not None else ""])
    for cell in overview[1]:
        cell.font = Font(bold=True)

    detail = workbook.create_sheet("明细")
    flat_rows = [flatten_row(row) for row in detail_rows]
    headers = list(dict.fromkeys(key for row in flat_rows for key in row))
    if headers:
        detail.append(headers)
        for cell in detail[1]:
            cell.font = Font(bold=True)
        for row in flat_rows:
            detail.append([row.get(header, "") for header in headers])
    else:
        detail.append(["message"])
        detail.append(["无明细数据"])

    by_adv = workbook.create_sheet("账户汇总")
    advertiser_rows = summary.get("by_advertiser") or []
    adv_headers = [
        "advertiser_id",
        "advertiser_name",
        "row_count",
        "spend",
        "cash_spend",
        "voucher_spend",
        "impressions",
        "engagements",
        "clicks",
        "conversion",
        "native_growth_ad_revenue_value_d0",
        "native_growth_ad_revenue_roas_d0",
        "purchase_value",
        "ctr",
        "cpc",
        "cpa",
    ]
    by_adv.append(adv_headers)
    for cell in by_adv[1]:
        cell.font = Font(bold=True)
    for row in advertiser_rows:
        by_adv.append([row.get(header, "") for header in adv_headers])

    fetch_sheet = workbook.create_sheet("拉取账户")
    fetch_headers = ["advertiser_id", "advertiser_name", "row_count"]
    fetch_sheet.append(fetch_headers)
    for cell in fetch_sheet[1]:
        cell.font = Font(bold=True)
    for row in account_summaries:
        fetch_sheet.append([row.get(header, "") for header in fetch_headers])

    error_sheet = workbook.create_sheet("错误")
    error_sheet.append(["advertiser_id", "advertiser_name", "error"])
    for cell in error_sheet[1]:
        cell.font = Font(bold=True)
    for row in errors:
        error_value = row.get("error")
        if not isinstance(error_value, str):
            error_value = str(error_value)
        error_sheet.append(
            [
                row.get("advertiser_id", ""),
                row.get("advertiser_name", ""),
                error_value,
            ]
        )

    workbook.save(path)
    return path


def default_output_stem(
    *,
    preset: str,
    start_date: str | None,
    end_date: str | None,
    query_lifetime: bool,
) -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if query_lifetime:
        return f"{stamp}_report_lifetime"
    start = start_date or "start"
    end = end_date or "end"
    safe_preset = "".join(
        ch if ch.isalnum() or ch in "-_" else "_" for ch in preset
    )
    return f"{stamp}_report_{safe_preset}_{start}_to_{end}"
